# from flask import Flask, request, jsonify, render_template, send_from_directory, send_file
# import os
# import importlib.util
# from bs4 import BeautifulSoup
# import io
# import pandas as pd
# from openpyxl import Workbook
# import re


# app = Flask(__name__)

# @app.route('/')
# def index():
#     return render_template('index.html')

# @app.route('/calculations/<filename>.html')
# def serve_calculation_html(filename):
#     """Serve calculation HTML files from calculations folder"""
#     try:
#         return send_from_directory('calculations', f"{filename}.html")
#     except FileNotFoundError:
#         return f"Calculation file '{filename}' not found", 404

# @app.route('/static/<path:filename>')
# def serve_static(filename):
#     """Serve static files"""
#     try:
#         return send_from_directory('static', filename)
#     except FileNotFoundError:
#         return f"Static file '{filename}' not found", 404

# @app.route('/api/calculate', methods=['POST'])
# def calculate():
#     try:
#         data = request.json
#         calc_type = data.get('calculation_type')
#         grid_data = data.get('data', [])
#         selected_columns = data.get('columns', [])
#         headers = data.get('headers', [])
#         additional_data = data.get('additional_data', {})

#         # Dynamic import of calculation module
#         calc_file = f"calculations/{calc_type}.py"
        
#         if not os.path.exists(calc_file):
#             return jsonify({
#                 'success': False,
#                 'error': f'Calculation type {calc_type} not found'
#             })

#         # Import the calculation module dynamically
#         spec = importlib.util.spec_from_file_location(calc_type, calc_file)
#         calc_module = importlib.util.module_from_spec(spec)
#         spec.loader.exec_module(calc_module)

#         # Call the calculate function
#         if hasattr(calc_module, 'calculate'):
#             result = calc_module.calculate(grid_data, selected_columns, headers, additional_data)
#         else:
#             return jsonify({
#                 'success': False,
#                 'error': f'No calculate function found in {calc_type}.py'
#             })
        
#         return jsonify({
#             'success': True,
#             'result': result
#         })

#     except Exception as e:
#         return jsonify({
#             'success': False,
#             'error': str(e)
#         })
    


# @app.route('/api/export', methods=['POST'])
# def export_all():
#     """
#     Improved export endpoint:
#     - Expects JSON { history: [ { type: "Label", output_html: "<table>...</table>" }, ... ] }
#     - For each history entry parses any <table> elements and writes them to Excel sheets.
#     - If no tables found for an entry, falls back to a text-based single-column sheet.
#     """
#     history = request.json.get('history', [])
#     if not history:
#         return jsonify({'error': 'No results to export'}), 400

#     output = io.BytesIO()
#     with pd.ExcelWriter(output, engine='openpyxl') as writer:
#         sheet_idx = 0
#         for entry in history:
#             # sheet base name (trim to Excel sheet limit)
#             base_name = str(entry.get('type', 'Sheet'))[:31] or f"Sheet{sheet_idx+1}"
#             html = entry.get('output_html', '') or ''

#             # Parse HTML for tables
#             soup = BeautifulSoup(html, 'html.parser')
#             tables = soup.find_all('table')

#             if tables:
#                 for t_i, table in enumerate(tables):
#                     # Extract rows
#                     rows = []
#                     for tr in table.find_all('tr'):
#                         # Prefer th if present in this row, otherwise td
#                         ths = tr.find_all('th')
#                         if ths:
#                             cells = [th.get_text(strip=True) for th in ths]
#                         else:
#                             tds = tr.find_all('td')
#                             cells = [td.get_text(strip=True) for td in tds]
#                         # Only append if there's any cell text (skip empty rows)
#                         if any(cell != '' for cell in cells):
#                             rows.append(cells)

#                     if not rows:
#                         continue

#                     # Normalize row lengths
#                     maxc = max(len(r) for r in rows)
#                     norm_rows = [r + [''] * (maxc - len(r)) for r in rows]

#                     # Decide DataFrame: if first row looks like header (has th in thead or the first tr had th),
#                     # use it as header; otherwise write all rows as-is.
#                     use_header = False
#                     # check for thead or first tr containing <th>
#                     if table.find('thead') is not None:
#                         use_header = True
#                     else:
#                         first_tr = table.find('tr')
#                         if first_tr and first_tr.find_all('th'):
#                             use_header = True

#                     if use_header and len(norm_rows) > 1:
#                         df = pd.DataFrame(norm_rows[1:], columns=norm_rows[0])
#                     else:
#                         # no header detected: create column names like Col1...ColN
#                         col_names = [f"Col{i+1}" for i in range(maxc)]
#                         df = pd.DataFrame(norm_rows, columns=col_names)

#                     # Write to a sheet; if multiple tables per entry, append suffix
#                     sheet_title = base_name if (t_i == 0) else f"{base_name}_{t_i+1}"
#                     sheet_title = sheet_title[:31]  # Excel sheet name limit
#                     try:
#                         df.to_excel(writer, sheet_name=sheet_title, index=False)
#                         sheet_idx += 1
#                     except Exception as e:
#                         # fallback: write as plain text sheet if DataFrame writing fails
#                         fallback_lines = [" | ".join(row) for row in norm_rows]
#                         fallback_df = pd.DataFrame({'Result': fallback_lines})
#                         fallback_title = (sheet_title + "_txt")[:31]
#                         fallback_df.to_excel(writer, sheet_name=fallback_title, index=False)
#                         sheet_idx += 1

#             else:
#                 # Fallback: no tables found -> export text content
#                 text = soup.get_text(separator='\n')
#                 lines = [ln for ln in (l.strip() for l in text.splitlines()) if ln]
#                 if not lines:
#                     continue
#                 df = pd.DataFrame({'Result': lines})
#                 try:
#                     df.to_excel(writer, sheet_name=base_name, index=False)
#                     sheet_idx += 1
#                 except Exception:
#                     # last-ditch fallback: small generic sheet
#                     small_df = pd.DataFrame({'Result': ["Export error writing sheet"]})
#                     small_df.to_excel(writer, sheet_name=(base_name)[:31], index=False)
#                     sheet_idx += 1

#         if sheet_idx == 0:
#             return jsonify({'error': 'No parsable results to export'}), 400

#     output.seek(0)
#     return send_file(
#         output,
#         as_attachment=True,
#         download_name='analysis_results.xlsx',
#         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )



# @app.route('/calculations/correlate.html')
# def correlate_popup():
#     # First popup: settings UI
#     return send_from_directory('calculations', 'correlate.html')

# @app.route('/calculations/corr_result.html')
# def corr_result_popup():
#     # Second popup: matrix display UI
#     return send_from_directory('calculations', 'corr_result.html')



# if __name__ == '__main__':
#     app.run(debug=True, host='0.0.0.0', port=5000)














# app.py
from flask import Flask, request, jsonify, render_template, send_from_directory, send_file
import os
import importlib.util
from bs4 import BeautifulSoup
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import re
import html as html_mod

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/calculations/<filename>.html')
def serve_calculation_html(filename):
    """Serve calculation HTML files from calculations folder"""
    try:
        return send_from_directory('calculations', f"{filename}.html")
    except FileNotFoundError:
        return f"Calculation file '{filename}' not found", 404

@app.route('/static/<path:filename>')
def serve_static(filename):
    """Serve static files"""
    try:
        return send_from_directory('static', filename)
    except FileNotFoundError:
        return f"Static file '{filename}' not found", 404

@app.route('/api/calculate', methods=['POST'])
def calculate():
    try:
        data = request.json
        calc_type = data.get('calculation_type')
        grid_data = data.get('data', [])
        selected_columns = data.get('columns', [])
        headers = data.get('headers', [])
        additional_data = data.get('additional_data', {})

        # Dynamic import of calculation module
        calc_file = f"calculations/{calc_type}.py"
        
        if not os.path.exists(calc_file):
            return jsonify({
                'success': False,
                'error': f'Calculation type {calc_type} not found'
            })

        # Import the calculation module dynamically
        spec = importlib.util.spec_from_file_location(calc_type, calc_file)
        calc_module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(calc_module)

        # Call the calculate function
        if hasattr(calc_module, 'calculate'):
            result = calc_module.calculate(grid_data, selected_columns, headers, additional_data)
        else:
            return jsonify({
                'success': False,
                'error': f'No calculate function found in {calc_type}.py'
            })
        
        return jsonify({
            'success': True,
            'result': result
        })

    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        })


# ---------- helper formatting functions for Excel export ----------
def _excel_autofit_ws(ws):
    """Auto-size columns based on cell contents (best-effort)."""
    for col in ws.columns:
        max_len = 0
        try:
            col_letter = get_column_letter(col[0].column)
        except Exception:
            continue
        for cell in col:
            if cell.value is None:
                continue
            val = str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        adjusted_width = (max_len + 2)
        ws.column_dimensions[col_letter].width = adjusted_width if adjusted_width < 60 else 60


def _apply_table_style(ws, start_row, start_col, end_row, end_col):
    """Apply borders, alignment and wrap_text to a rectangular area."""
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(start_row, end_row + 1):
    
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def _parse_mean_text_to_table(text):
    """
    Heuristic: parse text like shown by your mean function:
      Columns A
      Mean:
      2.1333
      Number of Observations:
      30
    Returns list of dict rows: [ { 'Variable': 'A', 'Mean': 2.1333, 'N': 30, 'Missing': ... }, ... ]
    """
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    rows = []
    i = 0
    while i < len(lines):
        line = lines[i]
        var_match = None
        # patterns like "Columns A" or "Columns X" or "Column A" or "A:" or "1: ColumnName"
        m = re.match(r'Columns?\s+(.+)', line, flags=re.I)
        if m:
            var = m.group(1).strip()
            mean_val = None
            n_val = None
            missing = None
            # look ahead a few lines
            j = i+1
            while j < min(i+8, len(lines)):
                l = lines[j]
                if l.lower().startswith('mean'):
                    # next non-empty line is the numeric value (or same line)
                    possible = lines[j+1] if j+1 < len(lines) else ''
                    if re.match(r'^-?\d+(\.\d+)?$', possible):
                        mean_val = float(possible)
                        j += 1
                if 'number of observations' in l.lower() or 'valid cases' in l.lower() or 'valid' in l.lower():
                    possible = lines[j+1] if j+1 < len(lines) else ''
                    if re.match(r'^\d+$', possible):
                        n_val = int(possible)
                        j += 1
                if 'missing' in l.lower():
                    possible = lines[j+1] if j+1 < len(lines) else ''
                    if re.match(r'^\d+$', possible):
                        missing = int(possible)
                        j += 1
                j += 1
            rows.append({'Variable': var, 'Mean': mean_val, 'N': n_val, 'Missing': missing})
            # advance i to j
            i = j
            continue

        # alternative format: lines like "1:Name" or "A: Name"
        m2 = re.match(r'^(\d+|[A-Za-z]{1,4})\s*:\s*(.+)$', line)
        if m2:
            var = m2.group(2).strip()
            # same look-ahead
            mean_val = None
            n_val = None
            j = i+1
            while j < min(i+8, len(lines)):
                l = lines[j]
                if l.lower().startswith('mean'):
                    possible = lines[j+1] if j+1 < len(lines) else ''
                    if re.match(r'^-?\d+(\.\d+)?$', possible):
                        mean_val = float(possible)
                        j += 1
                if 'number of observations' in l.lower() or 'valid cases' in l.lower():
                    possible = lines[j+1] if j+1 < len(lines) else ''
                    if re.match(r'^\d+$', possible):
                        n_val = int(possible)
                        j += 1
                j += 1
            rows.append({'Variable': var, 'Mean': mean_val, 'N': n_val, 'Missing': None})
            i = j
            continue

        i += 1

    return rows


def _parse_generic_block_text_to_kv_blocks(text):
    """
    Parse text into labelled blocks. Useful for correlation text output that
    wasn't returned as a <table>.
    Returns list of (label, [lines...])
    """
    lines = [ln.rstrip() for ln in text.splitlines()]
    blocks = []
    cur_label = None
    cur_lines = []
    for ln in lines:
        if not ln.strip():
            # blank line -> end current block
            if cur_label or cur_lines:
                blocks.append((cur_label if cur_label else '', [l for l in cur_lines if l.strip()]))
            cur_label = None
            cur_lines = []
            continue
        # if a line looks like a variable header (short), treat as label
        if re.match(r'^[A-Za-z0-9 _-]{1,30}$', ln) and len(ln.strip()) <= 20 and len(ln.split()) <= 4:
            # If we have a current block, close it
            if cur_label or cur_lines:
                blocks.append((cur_label if cur_label else '', [l for l in cur_lines if l.strip()]))
            cur_label = ln.strip()
            cur_lines = []
        else:
            cur_lines.append(ln)
    if cur_label or cur_lines:
        blocks.append((cur_label if cur_label else '', [l for l in cur_lines if l.strip()]))
    return blocks




def _descriptive_title(entry_type, idx, soup):
    # 1) prefer explicit title on export table
    t = soup.select_one('table.export-table')
    if t:
        for attr in ('data-title', 'aria-label', 'title'):
            v = (t.get(attr) or '').strip()
            if v:
                return f"{v} (analysis {idx+1})"
    # 2) visible headings
    for sel in ('h1','h2','h3','h4','.matrix-title','strong','b'):
        el = soup.select_one(sel)
        if el and el.get_text(strip=True):
            return f"{el.get_text(strip=True)} (analysis {idx+1})"
    # 3) heuristics for correlation variants
    if entry_type.lower() == 'correlation':
        txt = soup.get_text(separator=' ').lower()
        if 'partial' in txt or 'controlling-for' in txt:
            raw = soup.get_text(separator=' ')
            m = re.search(r'Partial[- ]Correlation\s*\(([^)]+)\)', raw, flags=re.I)
            extra = f" ({m.group(1).strip()})" if m else ''
            return f"Partial Correlation{extra} (analysis {idx+1})"
        if 'spearman' in txt:
            return f"Spearman Correlation (analysis {idx+1})"
        if 'kendall' in txt:
            return f"Kendall Correlation (analysis {idx+1})"
        if 'distance' in txt:
            if 'case-wise' in txt or 'casewise' in txt:
                return f"Distance Correlation (case-wise) (analysis {idx+1})"
            if 'variable-wise' in txt or 'variablewise' in txt:
                return f"Distance Correlation (variable-wise) (analysis {idx+1})"
            return f"Distance Correlation (analysis {idx+1})"
    return f"{entry_type.capitalize()} (analysis {idx+1})"






@app.route('/api/export', methods=['POST'])
def api_export():
    """
    Export 'history' array (frontend provided) to a single-sheet Excel workbook
    titled "Analysis Results". Each history item is appended with 5 blank rows
    between analyses. The function attempts to parse HTML tables and common
    text outputs (mean/correlation) into nicely formatted spreadsheet blocks.
    """
    try:
        payload = request.get_json(force=True) or {}
        history = payload.get('history', []) if payload else []

        # Keep only items that have non-empty output_html
        history = [h for h in history if h and h.get('output_html') and str(h.get('output_html')).strip()]
        if not history:
            return jsonify({"error": "No results to export"}), 400

        wb = Workbook()
        ws = wb.active
        ws.title = "Analysis Results"

        # Styles
        title_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        header_font = Font(bold=True)
        left_header_fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
        left_header_font = Font(bold=True)
        left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        current_row = 1

        for idx, item in enumerate(history):
            entry_type = str(item.get('type', 'analysis')).strip() or f"analysis_{idx+1}"
            # title = f"{entry_type.capitalize()} (analysis {idx+1})"
            html = item.get('output_html', '') or ''
            soup = BeautifulSoup(str(html), 'html.parser')
            title = _descriptive_title(entry_type, idx, soup)

            # Try to find tables first
            # tables = soup.find_all('table')

            exp_tables = soup.select('table.export-table')
            if exp_tables:
                tables = exp_tables
            else:
                tables = []
                # candidates = soup.find_all('table')
                # tables = [t for t in candidates if not is_hidden(t) and not is_grid(t)]
                # if not tables:
                #     tables = candidates
                for table in soup.find_all('table'):
                    style = (table.get('style') or '').lower()
                    classes = table.get('class', []) or []
                    # skip hidden tables
                    if 'export-table' in classes or 'display:none' in style:
                        continue
                    tables.append(table)


            if tables:
                # Use only first table per history item to keep structure consistent.
                # If multiple tables exist, append them sequentially.
                for t_i, table in enumerate(tables):
                    # title for this table (only write the entry title for the first table)
                    ws.cell(row=current_row, column=1, value=(title if t_i == 0 else f"{entry_type} (table {t_i+1})")).font = title_font
                    ws.cell(row=current_row, column=1).alignment = left_align
                    # if table has many columns, we won't merge — keep simple
                    current_row += 1

                    # collect rows and normalize columns
                    rows = []
                    has_th = False
                    for tr in table.find_all('tr'):
                        # if row has th use them as header cells
                        ths = tr.find_all('th')
                        if ths:
                            has_th = True
                            row_cells = [th.get_text(separator='\n').strip() for th in ths]
                        else:
                            tds = tr.find_all('td')
                            row_cells = [td.get_text(separator='\n').strip() for td in tds]
                        # convert multi-line cell text into single-line with ' ; ' separator
                        # row_cells = [' ; '.join([ln.strip() for ln in rc.splitlines() if ln.strip()]) for rc in row_cells]
                        row_cells = [re.sub(r'\n+', '\n', rc).strip() for rc in row_cells]
                        if any(rc != '' for rc in row_cells):
                            rows.append(row_cells)

                    if not rows:
                        ws.cell(row=current_row, column=1, value="(empty table)")
                        current_row += 1
                        continue

                    max_cols = max(len(r) for r in rows)
                    # normalize row lengths
                    norm_rows = [r + [''] * (max_cols - len(r)) for r in rows]

                    # decide header row: prefer explicit th/thead else use first row
                    header_row = None
                    data_rows = norm_rows
                    if has_th:
                        header_row = norm_rows[0]
                        data_rows = norm_rows[1:]
                    else:
                        # to keep spreadsheet readable, treat first row as header if it seems appropriate
                        header_row = norm_rows[0]
                        data_rows = norm_rows[1:]

                    start_data_row = current_row
                    # write header
                    for c_idx, h in enumerate(header_row, start=1):
                        cell = ws.cell(row=current_row, column=c_idx, value=h)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                    current_row += 1

                    # write data rows
                    for r in data_rows:
                        for c_idx in range(1, max_cols + 1):
                            val = r[c_idx - 1] if c_idx - 1 < len(r) else ''
                            cell = ws.cell(row=current_row, column=c_idx, value=(val if val != '' else None))
                            # left-most column get left-header styling for easy reading
                            if c_idx == 1:
                                cell.fill = left_header_fill
                                cell.font = left_header_font
                                cell.alignment = left_align
                            else:
                                cell.alignment = left_align
                        current_row += 1

                    # apply table border & autosize for the block
                    end_row = current_row - 1
                    end_col = max_cols
                    _apply_table_style(ws, start_data_row, 1, end_row, end_col)
                    _excel_autofit_ws(ws)

                    # small gap between multiple tables from same entry
                    current_row += 1

            else:
                # No table: try to parse common outputs
                text = soup.get_text(separator="\n").strip()
                text_lower = text.lower()

                written_any = False

                # 1) Mean-like output
                if 'mean' in text_lower and 'columns' in text_lower:
                    parsed = _parse_mean_text_to_table(text)
                    if parsed:
                        ws.cell(row=current_row, column=1, value=f"{title} - Mean").font = title_font
                        ws.cell(row=current_row, column=1).alignment = left_align
                        current_row += 1

                        headers = ['Variable', 'Mean', 'N', 'Missing']
                        for cidx, h in enumerate(headers, start=1):
                            cell = ws.cell(row=current_row, column=cidx, value=h)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                        current_row += 1

                        for row in parsed:
                            ws.cell(row=current_row, column=1, value=row.get('Variable'))
                            ws.cell(row=current_row, column=2, value=(round(row['Mean'], 4) if row.get('Mean') is not None else None))
                            ws.cell(row=current_row, column=3, value=row.get('N'))
                            ws.cell(row=current_row, column=4, value=row.get('Missing'))
                            current_row += 1

                        _apply_table_style(ws, current_row - len(parsed) - 1, 1, current_row - 1, 4)
                        _excel_autofit_ws(ws)
                        written_any = True

                # 2) Generic correlation-like or block output -> create 2-col block
                if not written_any:
                    blocks = _parse_generic_block_text_to_kv_blocks(text)
                    if blocks:
                        ws.cell(row=current_row, column=1, value=title).font = title_font
                        ws.cell(row=current_row, column=1).alignment = left_align
                        current_row += 1

                        start_block_row = current_row
                        for label, lines in blocks:
                            if label:
                                cell = ws.cell(row=current_row, column=1, value=label)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = left_align
                                current_row += 1
                            for ln in lines:
                                # try to split into kv by ":" otherwise entire text in first column
                                if ':' in ln:
                                    left, right = ln.split(':', 1)
                                    ws.cell(row=current_row, column=1, value=left.strip())
                                    ws.cell(row=current_row, column=2, value=right.strip())
                                else:
                                    ws.cell(row=current_row, column=1, value=ln)
                                current_row += 1
                            current_row += 1

                        _apply_table_style(ws, start_block_row, 1, current_row - 1, 2)
                        _excel_autofit_ws(ws)
                        written_any = True

                # 3) Last fallback: write entire text into a single cell
                if not written_any:
                    ws.cell(row=current_row, column=1, value=title).font = title_font
                    current_row += 1
                    cell = ws.cell(row=current_row, column=1, value=text if text else "(no readable output)")
                    cell.alignment = left_align
                    current_row += 1
                    _excel_autofit_ws(ws)

            # leave 5 blank rows between analyses
            current_row += 5

        # Save workbook to bytes and return
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        return send_file(
            out,
            as_attachment=True,
            download_name='analysis_results.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as ex:
        return jsonify({"error": f"Export failed: {str(ex)}"}), 500



@app.route('/calculations/correlate.html')
def correlate_popup():
    # First popup: settings UI
    return send_from_directory('calculations', 'correlate.html')

@app.route('/calculations/corr_result.html')
def corr_result_popup():
    # Second popup: matrix display UI
    return send_from_directory('calculations', 'corr_result.html')


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
